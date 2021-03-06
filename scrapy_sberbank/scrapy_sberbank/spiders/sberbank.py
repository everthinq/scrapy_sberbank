import json
import scrapy
from openpyxl import load_workbook


class sberbank(scrapy.Spider):
    name = 'sberbank'

    def start_requests(self):
        self.regions = [
            {"region": "0moskva i oblast.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D54.1458649223854%26llon%3D32.31010578124999%26rlat%3D57.246330825295644%26rlon%3D42.63725421875%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "10buryati.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D49.8726772079751%26llon%3D101.00520303117577%26rlat%3D53.2927916491364%26rlon%3D111.33235146867581%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "11vladimirskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D54.19817495451306%26llon%3D35.44486873144528%26rlat%3D57.29463170136643%26rlon%3D45.772017168945276%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "12volgogradskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D47.082073014087456%26llon%3D38.36161107812497%26rlat%3D50.702044550592404%26rlon%3D48.688759515624966%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "13vologodskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D58.740301557780214%26llon%3D38.886043781249924%26rlat%3D61.48044782107965%26rlon%3D49.213192218749946%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "14voronejskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D49.182423688040245%26llon%3D36.46508445312497%26rlat%3D52.65265234389037%26rlon%3D46.792232890624966%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "15dagestan.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D41.30596172386081%26llon%3D40.71268535937494%26rlat%3D45.31502494077085%26rlon%3D51.03983379687495%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "16evreiskaya AO.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D47.24182317114299%26llon%3D127.0798146317057%26rlat%3D50.850553709766594%26rlon%3D137.4069630692057%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "17zabaikal krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.65903089935094%26llon%3D111.21263103124994%26rlat%3D54.021524511475135%26rlon%3D121.53977946874993%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "18ivanovskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.43708586100162%26llon%3D36.80283178124993%26rlat%3D58.437942264749964%26rlon%3D47.129980218749914%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "19ingushetiya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D41.057791387732884%26llon%3D39.89100678124997%26rlat%3D45.08278313357833%26rlon%3D50.21815521874999%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "1spb.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D58.519877031746304%26llon%3D25.173775699423135%26rlat%3D61.27767317146125%26rlon%3D35.50092413692313%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "20irkutskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.329567%26llon%3D93.568636%26rlat%3D64.386718%26rlon%3D122.133087%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "21kabardino-balkariya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D41.461550756986%26llon%3D38.24469978124997%26rlat%3D45.460593309511275%26rlon%3D48.57184821874996%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "22kaliningrad.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D52.93229338781263%26llon%3D16.055369781250004%26rlat%3D56.12513415390239%26rlon%3D26.382518218749997%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "23kalmikiya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D44.48101609724257%26llon%3D40.16212678124993%26rlat%3D48.28045975276019%26rlon%3D50.489275218749924%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "24kalujskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D53.385133683127016%26llon%3D30.28161078124998%26rlat%3D56.543648697504295%26rlon%3D40.608759218749995%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "25kamchatskii krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.997803%26llon%3D153.492629%26rlat%3D65.202068%26rlon%3D175.133018%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "26karachaevo-cherkessk.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D42.58912239440853%26llon%3D37.37038307812499%26rlat%3D46.514760234863054%26rlon%3D47.697531515625%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "27kareliya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D60.552254%26llon%3D27.968226%26rlat%3D66.813919%26rlon%3D37.840162%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "28kemerovskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.939134%26llon%3D84.300816%26rlat%3D57.060105%26rlon%3D89.343541%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "29kirovskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.585720%26llon%3D45.130829%26rlat%3D61.392446%26rlon%3D54.513152%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "2adygeya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D42.42842506870239%26llon%3D35.08939478124998%26rlat%3D46.3646078494554%26rlon%3D45.41654321874999%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "30respublika komi.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D58.898030%26llon%3D44.782887%26rlat%3D68.827176%26rlon%3D66.447924%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "31kostromskaya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D56.90677592859415%26llon%3D38.40030104687491%26rlat%3D59.79265325713265%26rlon%3D48.72744948437493%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "32krasnodarskii krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D42.795434%26llon%3D36.007432%26rlat%3D46.981095%26rlon%3D42.136784%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "33krasnoyarskii krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.954346%26llon%3D76.470503%26rlat%3D82.164149%26rlon%3D116.584751%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "34kurganskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D53.914429655506%26llon%3D60.88728585937499%26rlat%3D57.03260631973592%26rlon%3D71.21443429687498%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "35kurskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D49.789530883525664%26llon%3D30.957772781249954%26rlat%3D53.21570514278531%26rlon%3D41.28492121874997%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "36leningradskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D58.357322%26llon%3D27.058467%26rlat%3D61.562151%26rlon%3D36.077553%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "37lipetskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.94167138845914%26llon%3D33.98620978124995%26rlat%3D54.283318383711666%26rlon%3D44.31335821874996%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "38magadanskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D58.345436%26llon%3D143.476012%26rlat%3D66.740906%26rlon%3D164.086362%26size%3D15%26page%3D1%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "39mariy-el.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.06538516452506%26llon%3D42.99377169628904%26rlat%3D58.09505316455226%26rlon%3D53.32092013378906%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "3altai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D48.96051859685835%26llon%3D81.85187978125%26rlat%3D52.44676460389284%26rlon%3D92.17902821874999%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "40mordoviya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D52.565729184978174%26llon%3D38.788099570406395%26rlat%3D55.78622891801186%26rlon%3D49.11524800790639%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "41moskovskaya oblast.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D54.136921824426814%26llon%3D32.977844378906205%26rlat%3D57.23807292171915%26rlon%3D43.3049928164062%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "42murmanskaya oblast.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D66.267049%26llon%3D27.190048%26rlat%3D70.111509%26rlon%3D42.476119%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "43nenetskiy AO.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D65.257558%26llon%3D42.637334%26rlat%3D70.980574%26rlon%3D65.762416%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "44nijegorodskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D54.284355%26llon%3D40.880713%26rlat%3D58.202134%26rlon%3D48.142067%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "45novgorodskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D56.81638271312735%26llon%3D27.56560041796875%26rlat%3D59.7093798115808%26rlon%3D37.89274885546874%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "46novosibirskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D53.110646%26llon%3D74.088719%26rlat%3D57.507152%26rlon%3D85.285440%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "47omskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D53.154824%26llon%3D69.400511%26rlat%3D58.658058%26rlon%3D76.560141%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "48orenburgskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.551651%26llon%3D50.162035%26rlat%3D54.479445%26rlon%3D62.192063%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "49orlovskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.722637968752046%26llon%3D31.284508796874974%26rlat%3D54.080446358130565%26rlon%3D41.611657234374995%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "4altai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.48410777659552%26llon%3D77.76194659374998%26rlat%3D53.859467319773216%26rlon%3D88.08909503124997%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "50penzenskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.718442976088724%26llon%3D39.69511401562496%26rlat%3D55.002433361461094%26rlon%3D50.02226245312495%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "51permskii krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.799485%26llon%3D51.320934%26rlat%3D61.927109%26rlon%3D60.483531%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "52primorskii krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D42.116688%26llon%3D129.861639%26rlat%3D48.613051%26rlon%3D139.694402%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "53pskovskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.519750%26llon%3D27.238824%26rlat%3D59.184777%26rlon%3D31.193902%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "54bashkortostan.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.855891%26llon%3D52.708038%26rlat%3D56.782608%26rlon%3D60.090850%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "55rostovskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D46.029076%26llon%3D37.566913%26rlat%3D50.483093%26rlon%3D44.442043%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "56ryazanskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D52.93191389775754%26llon%3D35.786756781249935%26rlat%3D56.124783357102466%26rlon%3D46.11390521874993%26size%3D9999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "57samarskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.756946371869645%26llon%3D45.889714470078005%26rlat%3D55.03806502565503%26rlon%3D56.21686290757802%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "58saratovskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D49.59004391934662%26llon%3D41.11797800488277%26rlat%3D53.03073125740551%26rlon%3D51.44512644238276%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "59sahalinskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D45.98436195346691%26llon%3D137.9203100650895%26rlat%3D49.68089930144815%26rlon%3D148.2474585025895%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "5amursk.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.71989473464008%26llon%3D122.89407962499992%26rlat%3D55.00377686572841%26rlon%3D133.22122806249996%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "60sverdlovskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D56.059923%26llon%3D56.653005%26rlat%3D62.522584%26rlon%3D66.694508%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "61severnaya osetiya-alaniya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D41.04650982565164%26llon%3D39.098458781249946%26rlat%3D45.07222406053849%26rlon%3D49.42560721874994%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "62smolenskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D53.127404%26llon%3D30.991003%26rlat%3D56.280876%26rlon%3D35.367785%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "63stavropolskiy krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D42.929580591988504%26llon%3D37.983192875000015%26rlat%3D46.832786690041516%26rlon%3D48.3103413125%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "64tambovskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D50.88355055599522%26llon%3D36.42867478125%26rlat%3D54.22949016943764%26rlon%3D46.75582321874999%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "65tatarstan.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D53.617116249743674%26llon%3D45.02234112499999%26rlat%3D56.75798020624351%26rlon%3D55.34948956249998%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "66tverskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.381009%26llon%3D30.680696%26rlat%3D59.340496%26rlon%3D37.700959%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "67tomskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.595405%26llon%3D73.821334%26rlat%3D61.537085%26rlon%3D89.663618%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "68tulskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D52.23370552662013%26llon%3D32.41211878124997%26rlat%3D55.479158958082444%26rlon%3D42.739267218749994%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "69respublika tyva.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D49.7146030025366%26llon%3D86.70362245969196%26rlat%3D53.146232692979424%26rlon%3D97.030770897192%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "6arhangelsk.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D62.38966094730467%26llon%3D38.173086781249985%26rlat%3D64.832711439315%26rlon%3D48.50023521874998%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "70tymenskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D54.593145%26llon%3D63.997290%26rlat%3D60.436026%26rlon%3D75.898642%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "71udmurtskaya resp.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D55.749662420601304%26llon%3D48.369481765624954%26rlat%3D58.72620548698233%26rlon%3D58.69663020312495%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "72ulyanoskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D52.02891944365011%26llon%3D42.54183573437499%26rlat%3D55.28971656242068%26rlon%3D52.868984171874956%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "73habarovskii krai.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D46.073638%26llon%3D128.659848%26rlat%3D63.075131%26rlon%3D149.489925%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "74resp hakasiya.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D52.304546467798524%26llon%3D85.32676549999998%26rlat%3D55.544683691644714%26rlon%3D95.65391393749998%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "75hanti-mansiyskii AO.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D58.482336%26llon%3D56.538171%26rlat%3D66.076269%26rlon%3D87.142580%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "76chelyabinskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.839947%26llon%3D55.888864%26rlat%3D56.792541%26rlon%3D64.1066374%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "77chechen resp.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D40.867567975804505%26llon%3D40.990586921874964%26rlat%3D44.90472321230894%26rlon%3D51.317735359374986%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "78chuvashskaya resp.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D53.66069436775909%26llon%3D41.96724609374994%26rlat%3D56.798237560400416%26rlon%3D52.29439453124996%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "79chukotskii AO.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D60.471783%26llon%3D151.873547%26rlat%3D70.381637%26rlon%3D179.034944%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "7astrahan.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D45.21356311616888%26llon%3D38.2652410506498%26rlat%3D48.96314849842853%26rlon%3D48.59238948814981%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "80resp saha.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D54.614685%26llon%3D102.197779%26rlat%3D77.847698%26rlon%3D155.779758%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "81yamalo-neneckiy AO.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D61.536681%26llon%3D58.730880%26rlat%3D74.816602%26rlon%3D85.454432%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "82yaroslavskaya obl.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D56.26482683967734%26llon%3D34.023274596679656%26rlat%3D59.20113172878819%26rlon%3D44.35042303417968%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "8belgorod.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D49.53256127579865%26llon%3D33.436010499999966%26rlat%3D52.97742381348462%26rlon%3D43.76315893749996%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"},
            {"region": "9bryansk.txt",
             "GET_request": "https://www.sberbank.ru/portalserver/proxy/?pipe=branchesPipe&url=http%3A%2F%2Flocalhost%2Foib-rs%2FbyBounds%2Fentities%3Fllat%3D51.88081410823111%26llon%3D30.4472678756068%26rlat%3D55.152685468391915%26rlon%3D40.77441631310679%26size%3D99999%26page%3D0%26cbLat%3D46.835967%26cbLon%3D29.606576%26filter%255Btype%255D%255B%255D%3Dfilial%26filter%255Bflags%255D%255BforPrivate%255D%3D1"}
        ]

        self.cell_value = '2'
        self.workbook = load_workbook('1. PJSC Sberbank Russia.xlsx')
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

        for region in self.regions:
            yield scrapy.Request(
                url=region['GET_request'],
                callback=self.parse_json
            )
            # break

    def parse_json(self, response):
        if (response.text):
            JSON = json.loads(response.text)
            # print(JSON)

            for branch in JSON:
                branch_name = branch['name']
                address = branch['postAddress']
                lat = branch['coordinates']['latitude']
                lng = branch['coordinates']['longitude']

                print('Writing --', 'name:', branch_name, 'address:', address, 'lat:', lat, 'lng:', lng)

                self.cell_value = str(self.cell_value)
                self.worksheet['B' + self.cell_value] = 'PJSC Sberbank Russia'
                self.worksheet['C' + self.cell_value] = branch_name
                self.worksheet['D' + self.cell_value] = address
                self.worksheet['G' + self.cell_value] = 'Russia'
                self.worksheet['H' + self.cell_value] = 'RU'
                self.worksheet['M' + self.cell_value] = lat
                self.worksheet['N' + self.cell_value] = lng
                self.worksheet['O' + self.cell_value] = 'Address'
                self.worksheet['R' + self.cell_value] = 'Bank website'
                self.cell_value = int(self.cell_value) + 1

        self.workbook.save('1. PJSC Sberbank Russia.xlsx')
